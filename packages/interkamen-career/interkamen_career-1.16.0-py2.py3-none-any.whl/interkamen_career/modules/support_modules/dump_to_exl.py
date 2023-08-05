#!/usr/bin/env python3
"""
Dump data to xlsx file

.dump_drill_pass()
.dump_ktu()
.dump_salary()
"""

import os
import time
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from .standart_functions import BasicFunctionsS as BasF_S


class DumpToExl(BasF_S):
    """Dump data to xlsx file."""

    __slots__ = ()

    months = [
        '', 'января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
        'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря'
    ]

    @classmethod
    def _create_pass_name(cls, passport) -> str:
        """Create passport name."""
        pass_name = ("{}-{}-{} {}"
                     .format(passport.params.year,
                             passport.params.month,
                             passport.params.day,
                             int(passport.params.number)))
        return pass_name

    @classmethod
    def _normilise_barehole(cls, bareholes):
        """Normalise 5+ meters bareholes to 5 meters."""
        count = 0
        new_bareholes = {}
        for key in bareholes:
            if key >= 5:
                count += bareholes[key]
            else:
                new_bareholes[key] = bareholes[key]
        if count:
            new_bareholes[5] = count
        return new_bareholes

    def _dump_real_salary(self, report: 'MainReport'):
        """Dumpp real salary to exel blanc."""
        real_salary_path = (
            super().get_root_path().parent / 'Documents' / 'Табеля'
        )
        real_salary_blanc_path = (
            super().get_root_path() / 'exl_blancs' / 'real_salary.xlsx'
        )
        workbook = load_workbook(real_salary_blanc_path)
        worksheet = workbook.active
        worksheet['C1'] = report.status['date']

        salary = report.workers_showing['факт']['зарплата']
        worker_number = 1
        for worker in salary:
            row_number = 4 + worker_number
            worksheet['B' + str(row_number)] = worker_number
            worksheet['C' + str(row_number)] = super().make_name_short(worker)
            worksheet['D' + str(row_number)] = salary[worker]
            worker_number += 1

        name = f"{report.status['date']} {report.status['shift']} на руки"
        pass_name = real_salary_path.joinpath(name).with_suffix('.xlsx')
        workbook.save(pass_name)

    def _fill_salary(self, workbook, report, user):
        """Fill salary to exel."""
        worksheet = workbook.active
        brigadiers_path = super().get_root_path() / 'data' / 'brigadiers'
        brigadiers = super().load_data(
            data_path=brigadiers_path,
            user=user,
        )
        worksheet['H28'] = report.totall
        ktu = report.workers_showing['бух.']['КТУ']
        hours = report.workers_showing['бух.']['часы']
        salary = report.workers_showing['бух.']['зарплата']

        worker_number = 1
        for worker in ktu:
            row_number = 30 + worker_number
            worksheet['B' + str(row_number)] = worker_number
            worksheet['C' + str(row_number)] = super().make_name_short(worker)
            worksheet['E' + str(row_number)] = hours[worker]
            worksheet['H' + str(row_number)] = ktu[worker]
            worksheet['K' + str(row_number)] = salary[worker]
            addition = 0
            if worker in brigadiers:
                addition = 0.15
            worksheet['N' + str(row_number)] = addition
            worksheet['Q' + str(row_number)] = (
                salary[worker] + salary[worker] * addition)
            worker_number += 1

    def dump_drill_pass(self, passport, negab=None):
        """Dump drill passport data to blanc exl file."""
        blanc_drill_path = (
            super().get_root_path() / 'exl_blancs' / 'drill_passport.xlsx')
        drill_pass_path = (
            super().get_root_path().parent / 'Documents' / 'Буровые_паспорта')
        workbook = load_workbook(blanc_drill_path)
        worksheet = workbook.active
        if negab:
            img = Image(
                super().get_root_path() / 'exl_blancs' / 'scheme_ng.png')
            worksheet['F4'] = 'колличество негабаритов:'
            worksheet['K4'] = int(negab)
        else:
            img = Image(super().get_root_path() / 'exl_blancs' / 'scheme.png')

        worksheet.add_image(img, 'A29')
        worksheet['K1'] = int(passport.params.number)  # Passport number.
        worksheet['J5'] = str(passport.params.day)  # Day.
        worksheet['K5'] = self.months[int(passport.params.month)]  # Month.
        worksheet['M5'] = str(passport.params.year)  # Year.
        worksheet['Q6'] = str(passport.params.horizond)  # Horizond.
        worksheet['F9'] = float(passport.params.pownder)  # Pownder.
        worksheet['K9'] = int(passport.params.d_sh)  # D_SH.
        worksheet['P9'] = int(passport.params.detonators)  # Detonators.
        # Bareholes.
        row_number = 15
        norm_bareholes = self._normilise_barehole(passport.bareholes)
        for length in norm_bareholes:
            worksheet['G' + str(row_number)] = length
            worksheet['D' + str(row_number)] = int(norm_bareholes[length])
            row_number += 1
        # Volume
        volume = round(float(passport.params.pownder) * 5 +
                       float(passport.params.d_sh) / 10, 1)
        worksheet['K27'] = volume
        # Block params.
        height = float(passport.params.block_height)
        worksheet['H25'] = height
        depth = float(passport.params.block_depth)
        worksheet['P25'] = depth
        worksheet['L25'] = round(volume / height / depth, 1)
        worksheet['M8'] = round((worksheet['L25'].value - 0.4) / int(round(
            (passport.params.block_width - 0.4) / 0.35, 0)), 3) * 1000
        # Master.
        master = super().make_name_short(str(passport.params.master))
        worksheet['J47'] = master
        # Save file.
        pass_name = self._create_pass_name(passport)
        pass_name = drill_pass_path.joinpath(pass_name).with_suffix('.xlsx')
        workbook.save(pass_name)
        print("\nФайл сохранен:\n", str(pass_name))

    def dump_ktu(self, report):
        """Dump KTU data to blanc exl file."""
        ktu_path = super().get_root_path().parent / 'Documents' / 'КТУ'
        blanc_ktu_path = super().get_root_path() / 'exl_blancs' / 'ktu.xlsx'
        ktu = report.workers_showing['бух.']['КТУ']
        hours = report.workers_showing['бух.']['часы']
        year = report.status['date'].split('-')[0]
        month = (
            self.months[int(report.status['date'].split('-')[1][:2])][:-1]+'е'
        )
        shift = report.status['shift']
        brig_list = {
            'Смена 1': 'Бригадой №1',
            'Смена 2': 'Бригадой №2'
        }
        brig = brig_list[shift]

        workbook = load_workbook(blanc_ktu_path)
        worksheet = workbook.active
        worksheet['C4'] = brig
        worksheet['C5'] = month
        worksheet['D5'] = year
        worker_number = 1
        for worker in ktu:
            row_number = 7 + worker_number
            worksheet['A' + str(row_number)] = worker_number
            worksheet['B' + str(row_number)] = super().make_name_short(worker)
            worksheet['C' + str(row_number)] = hours[worker]
            worksheet['D' + str(row_number)] = ktu[worker]
            worker_number += 1
        # Save file.
        pass_name = '-'.join([
            year, report.status['date'].split('-')[1][:2], shift])
        pass_name = ktu_path.joinpath(pass_name).with_suffix('.xlsx')
        workbook.save(pass_name)
        print("\nФайл сохранен:\n", str(pass_name))
        time.sleep(3)

    def dump_salary(self, report: 'MainReport', user):
        """Dump Salary to exists exel tabel."""
        salary_path = super().get_root_path().parent / 'Documents' / 'Табеля'
        name = report.status['date'] + ' ' + report.status['shift']
        find = None
        for file in os.listdir(salary_path):
            if name in file:
                find = name
                break
        if find:
            file_path = salary_path.joinpath(find).with_suffix('.xlsx')
            workbook = load_workbook(file_path)
            self._fill_salary(workbook, report, user)
            workbook.save(file_path)
            self._dump_real_salary(report)
            print("\nФайл сохранен:\n", salary_path.joinpath(find))
        else:
            print("Табель не найден в папке 'Табеля'")
        time.sleep(3)
