import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet

from performance import SHEET_NAME


def plt_export_chart(excel_file, output):
    excel = load_workbook(excel_file)
    table = excel[SHEET_NAME]
    assert isinstance(table, Worksheet)
    rows = table.max_row

    datax = []
    datay = []
    i = 1
    for row in range(2, rows + 1):
        datax.append(table.cell(row, 1).value)
        datay.append(table.cell(row, 2).value)
        i += 1

    plt.title("Total Memory Summary")
    plt.ylabel("Memory(KB)")
    plt.xlabel("Time")
    plt.plot(datax, datay, color='Red')
    plt.savefig(output)
    plt.close('all')


def plt_export_history(pkg:str, histories: list, output: str):
    data_avg = []
    data_max = []
    datax = []

    for history in histories:
        if history.pkg != pkg:
            continue
        if not history.success:
            continue
        data_avg.append(history.avg)
        data_max.append(history.max)
        datax.append(history.date)

    plt.title("History Memory Summary")
    plt.ylabel("Memory(KB)")
    plt.xlabel("Date")
    plt.plot(datax, data_max, color='Red')
    plt.plot(datax, data_avg, color='Blue')
    plt.savefig(output)
    plt.close('all')
